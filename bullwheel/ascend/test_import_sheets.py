# Copyright (c) 2026 Barrie's Ski and Sports
# All Rights Reserved
# Unauthorized copying or distribution of this file is prohibited.

"""Unit tests for build_import_sheet.

These pin the property that matters to Ascend's "Import from Excel": the finished
sheet's used range must end at the last real data row. Ascend imports every row in
the used range, so a sample row the template ships with — left behind by blanking a
cell's value instead of deleting the row — arrives as a phantom order line.
"""

import io
import zipfile

import frappe
import openpyxl
from frappe.tests import UnitTestCase

from bullwheel.ascend.import_sheets import build_import_sheet

PURCHASE_ORDER_TEMPLATE = "ascend_template_purchase_order.xlsx"
VENDOR_PRODUCTS_TEMPLATE = "ascend_template_vendor_products.xlsx"


def template_path(filename):
	"""Absolute path to an Ascend import template shipped with the app."""
	return frappe.get_app_path("bullwheel", "ascend", "import_templates", filename)


def worksheet_of(sheet_bytes):
	"""Reopen a built sheet and return its active worksheet, so assertions run against the
	file as a consumer would actually read it rather than the in-memory workbook."""
	return openpyxl.load_workbook(io.BytesIO(sheet_bytes)).active


def sheet_dimension(sheet_bytes):
	"""The <dimension ref="..."> the saved worksheet declares. Read straight out of the xlsx
	XML because that declared used range — not openpyxl's view of it — is what a consumer
	reading the file through OLEDB/Excel sees."""
	archive = zipfile.ZipFile(io.BytesIO(sheet_bytes))
	worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
	start = worksheet_xml.index('<dimension ref="') + len('<dimension ref="')
	return worksheet_xml[start:worksheet_xml.index('"', start)]


class UnitTestBuildImportSheet(UnitTestCase):

	def test_purchase_order_sheet_ends_at_the_last_item(self):
		"""Three order items produce exactly three data rows — the purchase order template's 64
		sample rows must not survive into the export, or Ascend imports 64 blank order lines."""
		rows = [
			{"Identifier": f"PART-{index}", "Description": f"Item {index}", "Qty": 2, "Cost": 1.5, "Comments": None}
			for index in range(3)
		]

		sheet_bytes = build_import_sheet(template_path(PURCHASE_ORDER_TEMPLATE), rows)

		self.assertEqual(sheet_dimension(sheet_bytes), "A1:E4")
		self.assertEqual(worksheet_of(sheet_bytes).max_row, 4)

	def test_empty_row_list_leaves_only_the_header(self):
		"""An export with no items is a header-only sheet, not 64 blank lines."""
		sheet_bytes = build_import_sheet(template_path(PURCHASE_ORDER_TEMPLATE), [])

		self.assertEqual(sheet_dimension(sheet_bytes), "A1:E1")
		self.assertEqual(worksheet_of(sheet_bytes).max_row, 1)

	def test_more_items_than_template_sample_rows(self):
		"""Row count is driven by the data, not the template: an order larger than the
		template's sample block still ends exactly at its last item."""
		rows = [{"Identifier": f"PART-{index}", "Qty": 1, "Cost": 1.0} for index in range(100)]

		sheet_bytes = build_import_sheet(template_path(PURCHASE_ORDER_TEMPLATE), rows)

		self.assertEqual(sheet_dimension(sheet_bytes), "A1:E101")
		self.assertEqual(worksheet_of(sheet_bytes).max_row, 101)

	def test_vendor_products_sheet_ends_at_the_last_item(self):
		"""The same guarantee holds for the Vendor Products template, which ships with its own
		block of sample rows."""
		sheet_bytes = build_import_sheet(
			template_path(VENDOR_PRODUCTS_TEMPLATE),
			[{"VPN": "PART-1", "Description": "Item", "Cost": 1.0}],
		)

		self.assertEqual(sheet_dimension(sheet_bytes), "A1:W2")
		self.assertEqual(worksheet_of(sheet_bytes).max_row, 2)

	def test_values_land_under_their_header(self):
		"""Each value is written under the column whose header matches its key, and an unknown
		key is ignored rather than shifting the remaining columns."""
		sheet_bytes = build_import_sheet(
			template_path(PURCHASE_ORDER_TEMPLATE),
			[{"Identifier": "PART-1", "Description": "Item", "Qty": 4, "Cost": 9.99,
			  "Comments": "note", "Nonexistent": "ignored"}],
		)

		worksheet = worksheet_of(sheet_bytes)
		self.assertEqual(
			[cell.value for cell in worksheet[2]],
			["PART-1", "Item", 4, 9.99, "note"],
		)

	def test_none_values_do_not_create_cells(self):
		"""A None value leaves its cell out entirely, so a column that is empty for every row
		never widens the used range with blanks."""
		sheet_bytes = build_import_sheet(
			template_path(PURCHASE_ORDER_TEMPLATE),
			[{"Identifier": "PART-1", "Description": "Item", "Qty": 1, "Cost": 1.0, "Comments": None}],
		)

		worksheet = worksheet_of(sheet_bytes)
		self.assertIsNone(worksheet.cell(row=2, column=5).value)
		self.assertEqual(worksheet.cell(row=2, column=1).value, "PART-1")
