# Copyright (c) 2026 Barrie's Ski and Sports
# All Rights Reserved
# Unauthorized copying or distribution of this file is prohibited.

import io

import frappe
import openpyxl


def build_import_sheet(template_path, rows):
	"""Fill an xlsx import template with `rows` and return the finished workbook as bytes.

	This handles only the file IO: it loads the template, deletes the sample rows the
	template ships with, and writes each row. `rows` is an ordered list of
	{template_column_header: value} dicts — one dict per worksheet row — where each value
	is placed under the column whose header matches the dict key. Keys with no matching
	header are ignored. Choosing the template and mapping a DocType's table onto these
	column headers is the caller's job. Returning bytes (rather than streaming) lets a
	caller reuse the result — e.g. bundle several sheets into one zip."""
	workbook = openpyxl.load_workbook(template_path)
	worksheet = workbook.active

	# Index each header to its 1-based column number.
	header_to_column = {
		cell.value: cell.column
		for cell in worksheet[1]
		if cell.value is not None
	}

	# Delete — never merely blank — the sample rows the template ships with. Clearing a
	# cell's value leaves the row itself in the saved file, so the finished sheet still
	# reports the template's original used range (the purchase order template carries 64
	# such rows, giving every export a used range of A1:E65 no matter how few items it
	# holds). Ascend's "Import from Excel" reads that used range, so leftover rows arrive
	# as blank order lines and inflate the imported order's item count.
	if worksheet.max_row > 1:
		worksheet.delete_rows(2, worksheet.max_row - 1)

	for row_number, row_values in enumerate(rows, start=2):
		for template_column, value in row_values.items():
			column_index = header_to_column.get(template_column)
			# A None value would still create a cell, re-widening the used range with a
			# blank; leaving the cell out entirely keeps the sheet exactly as wide as its
			# real data.
			if column_index is None or value is None:
				continue
			worksheet.cell(row=row_number, column=column_index, value=value)

	file_buffer = io.BytesIO()
	workbook.save(file_buffer)
	return file_buffer.getvalue()


def serve_file_download(filename, content, content_type=None):
	"""Hand `content` back to the browser as a file download through the Frappe response."""
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = content
	frappe.local.response.type = "download"
	if content_type:
		frappe.local.response.content_type = content_type


def generate_import_sheet(template_path, rows, filename):
	"""Build a single import sheet from `rows` and serve it as a browser download."""
	serve_file_download(filename, build_import_sheet(template_path, rows))
