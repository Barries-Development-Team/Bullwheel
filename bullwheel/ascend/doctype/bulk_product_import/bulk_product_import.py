# Copyright (c) 2026 Barrie's Ski and Sports
# All Rights Reserved
# Unauthorized copying or distribution of this file is prohibited.

import io

import frappe
import openpyxl
from frappe.model.document import Document

# Maps template column headers to New Product child table fieldnames.
# Columns absent from this map (ID, IsNonInventory, eCommerce, Color Code, JH QTY)
# are left blank because they have no equivalent field in the New Product DocType.
TEMPLATE_COLUMN_TO_FIELD = {
	"VPN": "vpn",
	"Category": "category",
	"Brand": "brand",
	"Description": "description",
	"Cost": "cost",
	"MSRP": "msrp",
	"UPC": "upc",
	"Color": "color",
	"Size": "size",
	"StyleNumber": "style_number",
	"StyleName": "style_name",
	"Year": "year",
	"Gender": "gender",
	"Season": "season",
	"MPN": "mpn",
	"CaseQty": "case_quantity",
	"CaseUPC": "case_upc",
	"CaseMSRP": "cast_msrp",
}

CASE_COLUMNS = {"CaseQty", "CaseUPC", "CaseMSRP"}


class BulkProductImport(Document):
	pass


@frappe.whitelist()
def generate_import_sheet(name):
	"""Build an Ascend Vendor Products import sheet from the saved child table rows and serve it as a download."""
	document = frappe.get_doc("Bulk Product Import", name)

	template_path = frappe.get_app_path(
		"bullwheel", "ascend", "import_templates", "Ascend Template_Vendor Products.xlsx"
	)

	workbook = openpyxl.load_workbook(template_path)
	worksheet = workbook.active

	# Index each header to its 1-based column number.
	header_to_column = {
		cell.value: cell.column
		for cell in worksheet[1]
		if cell.value is not None
	}

	# Erase any sample rows the template ships with.
	for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
		for cell in row:
			cell.value = None

	for row_number, product in enumerate(document.table_frll, start=2):
		for template_column, field_name in TEMPLATE_COLUMN_TO_FIELD.items():
			column_index = header_to_column.get(template_column)
			if column_index is None:
				continue
			if template_column in CASE_COLUMNS and not product.case:
				continue
			worksheet.cell(row=row_number, column=column_index, value=getattr(product, field_name, None))

	file_buffer = io.BytesIO()
	workbook.save(file_buffer)
	file_buffer.seek(0)

	frappe.local.response.filename = f"{name} - Ascend Vendor Product Import.xlsx"
	frappe.local.response.filecontent = file_buffer.read()
	frappe.local.response.type = "download"